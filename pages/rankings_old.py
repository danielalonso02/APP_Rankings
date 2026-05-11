#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rankings de Jugadoras
Plataforma profesional de análisis de rendimiento para jugadoras de Liga F
"""
import streamlit as st
from utils import login
from utils import util
import pandas as pd
import numpy as np
import os
import plotly.graph_objects as go
import plotly.express as px
from pathlib import Path
from datetime import datetime
from scipy.stats import zscore, percentileofscore
from utils.styles import inject_global_css
import io


# ==================== HELPER: EXCEL CON FORMATO ====================

def exportar_excel_bonito(df: pd.DataFrame, sheet_name: str = "Ranking") -> bytes:
    """
    Genera un Excel visualmente mejorado:
    - Cabecera azul marino con texto blanco en negrita
    - Filas alternas blanco / gris claro
    - Degradado rojo→amarillo→verde en columnas de Score, Percentil y Z-score
    - Bordes finos, ancho de columna automático y primera fila congelada
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill  = PatternFill("solid", fgColor="1E3A8A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    row_fill_alt = PatternFill("solid", fgColor="F1F5F9")
    row_fill_wht = PatternFill("solid", fgColor="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin_side    = Side(style="thin", color="CBD5E1")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side,  bottom=thin_side)

    # Cabecera
    headers = ["#"] + list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 22

    # Datos
    text_cols = {"Jugadora", "Equipo", "Posición", "Nota Global",
                 "Nota Ofensivo", "Nota Defensivo", "Nota Técnico", "Nota Físico"}
    for row_idx, (orig_idx, row) in enumerate(df.iterrows(), start=2):
        fill = row_fill_alt if row_idx % 2 == 0 else row_fill_wht
        idx_cell = ws.cell(row=row_idx, column=1,
                           value=int(orig_idx) if isinstance(orig_idx, (int, float)) else row_idx - 1)
        idx_cell.fill = fill
        idx_cell.font = Font(size=10, color="64748B")
        idx_cell.alignment = center_align
        idx_cell.border = thin_border

        for col_idx, (col_name, value) in enumerate(row.items(), start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = left_align if (col_name in text_cols or isinstance(value, str)) else center_align
            if isinstance(value, float):
                cell.number_format = "0.00"
                cell.font = Font(size=10)
            elif isinstance(value, int):
                cell.number_format = "0"
                cell.font = Font(size=10)
            else:
                cell.font = Font(size=10)

    # Formato condicional en columnas de score
    last_row = ws.max_row
    score_keywords = ("Score", "Percentil", "Z-score")
    for col_idx, col_name in enumerate(headers, start=1):
        if any(kw in str(col_name) for kw in score_keywords) and last_row > 2:
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="min",        start_color="EF4444",
                    mid_type="percentile",   mid_value=50, mid_color="F59E0B",
                    end_type="max",          end_color="10B981"
                )
            )

    # Ancho automático de columnas
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or ""))
              for r in range(2, last_row + 1)]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.set_page_config(
    page_title="Rankings | Liga F Analytics",
    page_icon=":material/leaderboard:",
    layout="wide",
    initial_sidebar_state="expanded"
)

inject_global_css()

# Sistema de login
login.generarLogin()

if "usuario" not in st.session_state:
    st.stop()

# Rutas de archivos auxiliares (pesos, parámetros, etc.)
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
ruta_parameters = os.path.join(BASE_DIR, "report_gen_opta", "parameters.xlsx")
folder_opta = Path(BASE_DIR) / "report_gen_opta"
folder_skillcorner = folder_opta / "SKILLCORNER"

# ==================== ENCABEZADO ====================
st.markdown("""
    <div style='text-align: center; margin-bottom: 2rem;'>
        <h1 style='color: #1e3a8a; margin-bottom: 0.5rem;'>
            🏆 Rankings de Jugadoras
        </h1>
        <p style='color: #64748b; font-size: 1.1rem;'>
            Clasificaciones y comparativas de rendimiento por métricas
        </p>
    </div>
""", unsafe_allow_html=True)

st.divider()

# ==================== CARGA DE DATOS ====================
if "df_rankings_archivos" not in st.session_state:
    st.session_state["df_rankings_archivos"] = {}

archivos_cargados = st.session_state["df_rankings_archivos"]

if archivos_cargados:
    col_info, col_cambiar = st.columns([3, 1])
    with col_info:
        nombres = list(archivos_cargados.keys())
        resumen = ", ".join(f"**{n}**" for n in nombres)
        total_filas = sum(len(df) for df in archivos_cargados.values())
        st.success(f"✅ {len(archivos_cargados)} archivo(s) cargado(s): {resumen} · {total_filas:,} jugadoras en total")
    with col_cambiar:
        if st.button("🔄 Cambiar archivos", use_container_width=True, key="cambiar_rankings"):
            st.session_state["df_rankings_archivos"] = {}
            st.rerun()
else:
    st.info("ℹ️ Sube uno o varios archivos Excel para comenzar. Cada archivo se tratará como una liga independiente.")
    uploaded_files = st.file_uploader(
        "📂 Archivos de datos de jugadoras (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="uploader_rankings"
    )
    if uploaded_files:
        nuevos = {}
        errores = []
        for uf in uploaded_files:
            try:
                import io as _io_upload
                df_tmp = pd.read_excel(_io_upload.BytesIO(uf.read()))
                for col in df_tmp.columns:
                    if df_tmp[col].dtype == object:
                        converted = df_tmp[col].astype(str).str.replace(",", ".", regex=False)
                        numeric = pd.to_numeric(converted, errors="coerce")
                        if numeric.notna().sum() > df_tmp[col].notna().sum() * 0.5:
                            df_tmp[col] = numeric
                nuevos[uf.name] = df_tmp
            except Exception as e:
                errores.append(f"{uf.name}: {e}")
        if errores:
            for err in errores:
                st.error(f"❌ {err}")
        if nuevos:
            st.session_state["df_rankings_archivos"] = nuevos
            st.rerun()
    st.stop()

def _competicion_de_nombre(nombre_archivo: str) -> str:
    """
    Extrae un identificador único de competición a partir del nombre del archivo.
    Patrón esperado: pais_division_YYYY_MM_DD.xlsx
    Devuelve 'pais_division' (las dos primeras partes) para distinguir
    argentina_1 de argentina_2, libertadores_0, etc.
    """
    partes = os.path.splitext(nombre_archivo)[0].split("_")
    if len(partes) >= 2:
        return f"{partes[0]}_{partes[1]}"
    return partes[0]

# df_full = unión de todos los archivos, con columna _liga para identificar origen
_dfs_con_liga = []
for _nombre_archivo, _df_liga in archivos_cargados.items():
    _df_tmp = _df_liga.copy()
    _df_tmp["_liga"] = _competicion_de_nombre(_nombre_archivo)
    _df_tmp["_archivo"] = _nombre_archivo
    _dfs_con_liga.append(_df_tmp)
df_full = pd.concat(_dfs_con_liga, ignore_index=True)

season = "24/25"

# Diccionario de agrupación de posiciones
POSITIONS_DICT = {
    "GK": "Portero", "LB": "Lateral", "RB": "Lateral",
    "DMF": "Mediocampista defensivo", "OF": "Mediocampista ofensivo",
    "AMF": "Mediocampista ofensivo", "LCB": "Defensa central", "RCB": "Defensa central",
    "LWF": "Extremo", "RWF": "Extremo", "LAMF": "Mediocampista ofensivo",
    "RAMF": "Mediocampista ofensivo", "LCMF": "Mediocampista ofensivo",
    "RCMF": "Mediocampista ofensivo", "CF": "Delantero", "CB": "Defensa central",
    "RW": "Extremo", "LDMF": "Mediocampista defensivo", "LW": "Extremo",
    "RDMF": "Mediocampista defensivo", "RWB": "Lateral", "LWB": "Lateral",
}

EXP_POSITION = {
    "GK": "Portero", "LB": "Lateral izquierdo", "RB": "Lateral derecho",
    "DMF": "Medio defensivo", "OF": "Banda",
    "AMF": "Medio ofensivo", "LCB": "Defensa central izquierdo", 
    "RCB": "Defensa central derecho", "LWF": "Extremo izquierdo", 
    "RWF": "Extremo derecho", "LAMF": "Medio ofensivo izquierdo",
    "RAMF": "Medio ofensivo derecho", "LCMF": "Medio izquierdo",
    "RCMF": "Medio derecho", "CF": "Delantero centro", "CB": "Defensa central",
    "RW": "Extremo derecho", "LDMF": "Medio defensivo izquierdo", 
    "LW": "Extremo izquierdo", "RDMF": "Medio defensivo derecho", 
    "RWB": "Carrilero derecho", "LWB": "Carrilero izquierdo",
}

# Creamos las etiquetas tipo "Portero (GK)"
MAP_LABELS = {k: f"{v} ({k})" for k, v in EXP_POSITION.items()}
# Diccionario inverso para saber que "Portero (GK)" es "GK"
INV_MAP_LABELS = {v: k for k, v in MAP_LABELS.items()}

grupos_con_cods = {}
for cod, grp in POSITIONS_DICT.items():
    grupos_con_cods.setdefault(grp, []).append(cod)

MAP_LABELS_GRUPO = {n: f"{n} ({', '.join(c)})" for n, c in grupos_con_cods.items()}
INV_MAP_LABELS_GRUPO = {v: k for k, v in MAP_LABELS_GRUPO.items()}

col_eq, col_vacio = st.columns([2, 1])
with col_eq:
    # Los equipos vienen del Excel cargado por el usuario
    lista_equipos = sorted(df_full["Equipo durante el período seleccionado"].dropna().unique().tolist()) if "Equipo durante el período seleccionado" in df_full.columns else []
    equipos_sel = st.multiselect(
        "Equipos",
        options=lista_equipos,
        default=[],
        placeholder="Todos los equipos",
        key="equipo_ranking"
    )


def get_grupo_posicion(cell):
    """Devuelve el grupo de la primera posición reconocida en la celda."""
    if pd.isna(cell):
        return None
    for pos in str(cell).split(","):
        grupo = POSITIONS_DICT.get(pos.strip())
        if grupo:
            return grupo
    return None

if "Posición específica" in df_full.columns:
    df_full["Grupo posición"] = df_full["Posición específica"].apply(get_grupo_posicion)
    
col4, col_edad_fil = st.columns([1.5, 1.5])

with col4:
    min_min_abs = int(df_full["Minutos jugados"].min()) if "Minutos jugados" in df_full.columns else 0
    max_min_abs = int(df_full["Minutos jugados"].max()) if "Minutos jugados" in df_full.columns else 3000
    rango_minutos = st.slider(
        "Rango de minutos jugados",
        min_value=min_min_abs,
        max_value=max_min_abs,
        value=(max(500, min_min_abs), max_min_abs),
        step=10,
        key="rango_minutos_ranking"
    )




    
if "Edad" in df_full.columns:
    edad_min_abs = int(df_full["Edad"].dropna().min())
    edad_max_abs = int(df_full["Edad"].dropna().max())
    with col_edad_fil:
        rango_edad = st.slider(
            "Rango de edad",
            min_value=edad_min_abs,
            max_value=edad_max_abs,
            value=(edad_min_abs, edad_max_abs),
            step=1,
            key="rango_edad"
        )
    edad_activo = rango_edad != (edad_min_abs, edad_max_abs)
else:
    rango_edad = None
    edad_activo = False

# Segunda fila de filtros: grupo de posición + edad
col_grupo, col5 = st.columns([2, 2])

with col_grupo:
    grupos_disponibles_raw = sorted(df_full["Grupo posición"].dropna().unique().tolist()) if "Grupo posición" in df_full.columns else []
    
    # Mostramos la etiqueta larga: "Defensa central (CB, LCB, RCB)"
    opciones_grupo_mostrar = [MAP_LABELS_GRUPO.get(g, g) for g in grupos_disponibles_raw]
    
    grupos_sel_labels = st.multiselect(
        "Grupo de posición",
        options=opciones_grupo_mostrar,
        default=[],
        placeholder="Todos los grupos",
        key="grupo_posicion_ranking"
    )
    
    # Traducimos a nombre limpio para buscar en el DF y para la lógica de Pesos
    grupos_sel = [INV_MAP_LABELS_GRUPO.get(label, label) for label in grupos_sel_labels]

with col5:
    if "Posición específica" in df_full.columns:
        # Extraer todos los valores únicos dividiendo por coma (hasta 3 posiciones por jugadora)
        lista_posiciones = sorted(set(
            pos.strip()
            for cell in df_full["Posición específica"].dropna()
            for pos in str(cell).split(",")
            if pos.strip()
        ))
        
        opciones_mostrar = [MAP_LABELS.get(cod, cod) for cod in lista_posiciones]
    else:
        lista_posiciones = []
    posiciones_sel = st.multiselect(
        "Posiciones",
        options=opciones_mostrar,
        default=[],
        placeholder="Todas las posiciones",
        key="posicion_ranking"
    )


def _nombre_descarga(sufijo: str = "", competicion: str = "") -> str:
    """Construye: ranking_<posicion>_<competicion>[_sufijo].xlsx"""
    pos_part = (
        "-".join(INV_MAP_LABELS.get(lbl, lbl) for lbl in posiciones_sel).replace(" ", "")
        if posiciones_sel else "todas"
    )
    comp = competicion if competicion else (
        _competicion_de_nombre(list(archivos_cargados.keys())[0])
        if len(archivos_cargados) == 1 else "varias_ligas"
    )
    partes = ["ranking", pos_part, comp]
    if sufijo:
        partes.append(sufijo)
    return "_".join(partes) + ".xlsx"


# Aplicar filtros
df_filtered = df_full.copy()

# Filtro de minutos por rango
if "Minutos jugados" in df_filtered.columns:
    df_filtered = df_filtered[
        df_filtered["Minutos jugados"].between(rango_minutos[0], rango_minutos[1])
    ]

if equipos_sel:
    df_filtered = df_filtered[df_filtered["Equipo durante el período seleccionado"].isin(equipos_sel)]

if posiciones_sel:
    posiciones_sel_codes = [INV_MAP_LABELS.get(label, label) for label in posiciones_sel]
    def tiene_posicion(cell):
        if pd.isna(cell):
            return False
        # Obtenemos los códigos de la celda (ej: "GK, LB") y limpiamos espacios
        codigos_en_celda = [p.strip() for p in str(cell).split(",")]
        # Buscamos si alguno de los códigos seleccionados está en la jugadora
        return any(code in codigos_en_celda for code in posiciones_sel_codes)
    
    df_filtered = df_filtered[df_filtered["Posición específica"].apply(tiene_posicion)]

if grupos_sel and "Grupo posición" in df_filtered.columns:
    df_filtered = df_filtered[df_filtered["Grupo posición"].isin(grupos_sel)]

if edad_activo and rango_edad is not None and "Edad" in df_filtered.columns:
    df_filtered = df_filtered[
        df_filtered["Edad"].between(rango_edad[0], rango_edad[1])
    ]

# Verificación de datos resultantes
if df_filtered.empty:
    st.warning("⚠️ No hay jugadoras que cumplan los filtros seleccionados.")
    st.stop()

st.divider()

# ==================== MAPEO DE COLUMNAS WYSCOUT ====================
# Las columnas de Wyscout ya vienen en /90 o como ratios.
# Solo necesitamos renombrarlas a los nombres internos que usa el resto del código.

def _to_num(series):
    """Convierte una serie a numérico, manejando coma decimal."""
    if pd.api.types.is_numeric_dtype(series):
        return series
    return pd.to_numeric(series.astype(str).str.replace(",", ".", regex=False), errors="coerce")

def add_per90_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Mapa: nombre interno → columna exacta del Excel Wyscout (todas ya en /90)
    rename_map = {
        "Goles /90":                     "Goles/90",
        "xG /90":                        "xG/90",
        "Asistencias /90":               "Asistencias/90",
        "xA /90":                        "xA/90",
        "Remates totales /90":           "Remates/90",
        "Pases clave /90":               "Jugadas claves/90",
        "Regates exitosos /90":          "Regates/90",
        "Centros exitosos /90":          "Centros/90",
        "Toques en área rival /90":      "Toques en el área de penalti/90",
        "Interceptaciones /90":          "Interceptaciones/90",
        "Entradas ganadas /90":          "Entradas/90",
        "Tiros interceptados /90":       "Tiros interceptados/90",
        "Faltas cometidas /90":          "Faltas/90",
        "Faltas recibidas /90":          "Faltas recibidas/90",
        "Pases exitosos /90":            "Pases/90",
        "Pases largos exitosos /90":     "Pases largos/90",
        "Pases en campo rival /90":      "Pases en el último tercio/90",
        "Carries progresivos /90":       "Carreras en progresión/90",
        "Segundas asistencias /90":      "Second assists/90",
        "Balones al hueco /90":          "Pases en profundidad/90",
        "Duelos totales /90":            "Duelos/90",
        "Duelos aéreos totales /90":     "Duelos aéreos en los 90",
        "Duelos ganados /90":            "Duelos ganados, %",
        "Duelos aéreos ganados /90":     "Duelos aéreos ganados, %",
        "Duelos defensivos /90":         "Duelos defensivos/90",
        "Duelos defensivos ganados /90": "Duelos defensivos ganados, %",
        "Recuperacies tras interceptación /90":         "Posesión conquistada después de una interceptación",
        "Recuperaciones tras entrada /90":            "Posesión conquistada después de una entrada",
        "Grandes ocasiones creadas /90": "Acciones de ataque exitosas/90",
        # Métricas presentes en pesos_scouting pero sin mapeo previo
        "Remates a puerta /90":          "Remates/90",
        "Despejes /90":                  "Tiros interceptados/90",
        "Recuperaciones /90":            "Acciones defensivas realizadas/90",
        "Bloqueos /90":                  "Tiros interceptados/90",
        "Duelos en suelo ganados /90":   "Duelos defensivos/90",
        "Pérdidas de balón /90":         "Acciones defensivas realizadas/90",
        "Ocasiones falladas /90":        "Acciones de ataque exitosas/90",
        "Faltas recibidas /90":          "Faltas recibidas/90",
    }

    # Renombrado con conversión numérica segura
    for nombre_interno, col_wyscout in rename_map.items():
        if col_wyscout in df.columns:
            df[nombre_interno] = _to_num(df[col_wyscout])

    return df

df_filtered = add_per90_cols(df_filtered)

# ── Mapa completo de métricas disponibles (nombre interno → columna en df) ──
# Esta es la fuente de verdad para el visualizador y la tabla multi-métrica.
# Es independiente del Excel de pesos: si añades una métrica en rename_map, 
# automáticamente aparece aquí.
RENAME_MAP_METRICAS = {
    "Goles /90":                            "Goles /90",
    "xG /90":                               "xG /90",
    "Asistencias /90":                      "Asistencias /90",
    "xA /90":                               "xA /90",
    "Remates totales /90":                  "Remates totales /90",
    "Pases clave /90":                      "Pases clave /90",
    "Regates exitosos /90":                 "Regates exitosos /90",
    "Centros exitosos /90":                 "Centros exitosos /90",
    "Toques en área rival /90":             "Toques en área rival /90",
    "Interceptaciones /90":                 "Interceptaciones /90",
    "Entradas ganadas /90":                 "Entradas ganadas /90",
    "Tiros interceptados /90":              "Tiros interceptados /90",
    "Faltas cometidas /90":                 "Faltas cometidas /90",
    "Faltas recibidas /90":                 "Faltas recibidas /90",
    "Pases exitosos /90":                   "Pases exitosos /90",
    "Pases largos exitosos /90":            "Pases largos exitosos /90",
    "Pases en campo rival /90":             "Pases en campo rival /90",
    "Carries progresivos /90":              "Carries progresivos /90",
    "Segundas asistencias /90":             "Segundas asistencias /90",
    "Balones al hueco /90":                 "Balones al hueco /90",
    "Duelos totales /90":                   "Duelos totales /90",
    "Duelos aéreos totales /90":            "Duelos aéreos totales /90",
    "Duelos ganados /90":                   "Duelos ganados /90",
    "Duelos aéreos ganados /90":            "Duelos aéreos ganados /90",
    "Duelos defensivos /90":                "Duelos defensivos /90",
    "Duelos defensivos ganados /90":        "Duelos defensivos ganados /90",
    "Recuperaciones tras interceptación /90": "Recuperacies tras interceptación /90",
    "Recuperaciones tras entrada /90":      "Recuperaciones tras entrada /90",
    "Grandes ocasiones creadas /90":        "Grandes ocasiones creadas /90",
    "Remates a puerta /90":                 "Remates a puerta /90",
    "Despejes /90":                         "Despejes /90",
    "Recuperaciones /90":                   "Recuperaciones /90",
    "Bloqueos /90":                         "Bloqueos /90",
    "Duelos en suelo ganados /90":          "Duelos en suelo ganados /90",
    "Pérdidas de balón /90":                "Pérdidas de balón /90",
    "Ocasiones falladas /90":               "Ocasiones falladas /90",
    "Faltas recibidas /90":                 "Faltas recibidas /90",
}

# ==================== CARGAR PESOS DESDE pesos_scouting.xlsx ====================

COL_MAP = {
    "ATAQUE":  "Ofensivo",
    "DEFENSA": "Defensivo",
    "TÉCNICA": "Técnico",
    "FÍSICO":  "Físico",
}

# ==================== CARGAR FUNCIONES Y CONFIGURACIÓN INICIAL ====================

# Ruta de archivos de pesos
ruta_pesos = folder_opta / "datoswyscout" / "pesos_scouting.xlsx"
ruta_pesos_macro = folder_opta / "datoswyscout" / "pesos_scouting_global.xlsx"

import json

# Ruta para guardar los favoritos
RUTA_FAVORITOS = Path(BASE_DIR)  / "config_pesos_favoritos.json"

def cargar_favoritos():
    if RUTA_FAVORITOS.exists():
        with open(RUTA_FAVORITOS, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_favorito(nombre, w_global, w_micro):
    favs = cargar_favoritos()
    favs[nombre] = {
        "global": w_global,
        "micro": w_micro
    }
    with open(RUTA_FAVORITOS, "w", encoding="utf-8") as f:
        json.dump(favs, f, indent=4, ensure_ascii=False)
        
def eliminar_favorito(nombre):
    favs = cargar_favoritos()
    if nombre in favs:
        del favs[nombre]
        with open(RUTA_FAVORITOS, "w", encoding="utf-8") as f:
            json.dump(favs, f, indent=4, ensure_ascii=False)
        return True
    return False

def aplicar_favorito(nombre):
    favs = cargar_favoritos()
    if nombre in favs:
        config = favs[nombre]
        # Aplicar Globales
        st.session_state["w_of_val"] = config["global"]["Ofensivo"]
        st.session_state["w_def_val"] = config["global"]["Defensivo"]
        st.session_state["w_tec_val"] = config["global"]["Técnico"]
        st.session_state["w_fis_val"] = config["global"]["Físico"]
        # Aplicar Micro (Sliders individuales)
        for key_slider, valor in config["micro"].items():
            st.session_state[key_slider] = valor
        st.success(f"Configuración '{nombre}' cargada correctamente.")

@st.cache_data
def cargar_pesos_macro_tabla(ruta_macro: str, posicion_buscada: str) -> dict:
    try:
        df_m = pd.read_excel(ruta_macro)
        df_m.columns = df_m.columns.str.strip()
        # Filtramos por la columna description
        fila = df_m[df_m["description"] == posicion_buscada].iloc[0]
        return {
            "Ofensivo":  float(fila["Ofensivo"]) / 100,
            "Defensivo": float(fila["Defensivo"]) / 100,
            "Técnico":   float(fila.get("Técnico", 0)) / 100,
            "Físico":    float(fila.get("Físico", 0)) / 100
        }
    except:
        return None

# Tabla de corrección de typos en nombres de variables del Excel
_TYPOS_VARIABLES = {
    "xA/ 90":           "xA /90",
    "xA 90/":           "xA /90",
    "xG /90 ":          "xG /90",
    "xG/ 90":           "xG /90",
    "Faltas recibidad /90": "Faltas recibidas /90",
}

@st.cache_data
def cargar_pesos_por_posicion(ruta: str, posicion: str) -> dict:
    try:
        df_p = pd.read_excel(ruta, sheet_name=posicion)
        df_p.columns = df_p.columns.str.strip()
        # Limpiar nombres: strip + corregir typos conocidos
        df_p["VARIABLE"] = (
            df_p["VARIABLE"]
            .str.strip()
            .replace(_TYPOS_VARIABLES)
        )
        df_p = df_p.set_index("VARIABLE")
        
        pesos = {cat_es: {} for cat_es in COL_MAP.values()}
        for col_xlsx, cat_es in COL_MAP.items():
            if col_xlsx not in df_p.columns: continue
            for variable, valor in df_p[col_xlsx].items():
                if pd.notna(valor) and valor != 0:
                    pesos[cat_es][variable] = float(valor)
        return pesos
    except:
        return None

# Inicialización de estados para evitar NameError y permitir sincronización
if "posicion_previa" not in st.session_state: st.session_state["posicion_previa"] = []
if "w_of_val" not in st.session_state: st.session_state["w_of_val"] = 0.25
if "w_def_val" not in st.session_state: st.session_state["w_def_val"] = 0.25
if "w_tec_val" not in st.session_state: st.session_state["w_tec_val"] = 0.25
if "w_fis_val" not in st.session_state: st.session_state["w_fis_val"] = 0.25

# FIX: Persistencia de selecciones por nombre (no por índice posicional)
# Cada tabla tiene su propio set de nombres seleccionados
for _key in ["sel_nombres_Global", "sel_nombres_Ofensivo", "sel_nombres_Defensivo",
             "sel_nombres_Técnico", "sel_nombres_Físico", "sel_nombres_ind", "sel_nombres_multi"]:
    if _key not in st.session_state:
        st.session_state[_key] = set()

# FIX: Persistencia de la pestaña activa para evitar saltos al inicio al cambiar checkboxes
if "tab_ranking_activa" not in st.session_state:
    st.session_state["tab_ranking_activa"] = 0

# --- LÓGICA DE SINCRONIZACIÓN AUTOMÁTICA (Filtro -> Sliders) ---
if grupos_sel != st.session_state["posicion_previa"]:
    if len(grupos_sel) == 1:
        grupo_activo = grupos_sel[0]
        
        # Mapeo: Nombre en el filtro -> Nombre en la columna 'description' del Excel Global
        map_macro = {
            "Delantero": "Delantero centro",
            "Mediocampista ofensivo": "Mediocampista ofensivo",
            "Mediocampista defensivo": "Mediocampista defensivo",
            "Defensa central": "Defensa central",
            "Lateral": "Lateral",
            "Extremo": "Extremo",
            "Portero": "Portero"
        }
        nombre_macro = map_macro.get(grupo_activo, grupo_activo)
        
        # 1. Cargar Macro y actualizar los KEYS de los sliders directamente
        m_vals = cargar_pesos_macro_tabla(str(ruta_pesos_macro), nombre_macro)
        if m_vals:
            st.session_state["w_of_val"] = m_vals["Ofensivo"]
            st.session_state["w_def_val"] = m_vals["Defensivo"]
            st.session_state["w_tec_val"] = m_vals["Técnico"]
            st.session_state["w_fis_val"] = m_vals["Físico"]

        # 2. Cargar Micro (Nivel 2)
        p_vals = cargar_pesos_por_posicion(str(ruta_pesos), grupo_activo)
        if p_vals:
            for cat, metricas in p_vals.items():
                for m, v in metricas.items():
                    st.session_state[f"slider_{cat}_{m}"] = float(v)
        
        st.session_state["posicion_previa"] = grupos_sel
        st.rerun()

# ==================== NIVEL 1: PESOS GLOBALES (MACRO) ====================
with st.expander("⚖️ NIVEL 1: Importancia de los Bloques"):
    
    # --- SECCIÓN DE FAVORITOS (CON BORRADO) ---
    st.markdown("<p style='font-weight: bold; color: #1e3a8a;'>⭐ Mis Configuraciones Guardadas</p>", unsafe_allow_html=True)
    
    # Ajustamos columnas: Selectbox, Cargar, Eliminar, Guardar
    f_col1, f_col2, f_col3, f_col4 = st.columns([2, 0.8, 0.8, 2])
    
    with f_col1:
        mis_favs = cargar_favoritos()
        nombre_sel = st.selectbox(
            "Cargar favorito:", 
            ["Seleccionar..."] + list(mis_favs.keys()),
            label_visibility="collapsed"
        )
    
    with f_col2:
        # Botón Cargar
        if st.button("📥 Cargar", use_container_width=True, disabled=nombre_sel=="Seleccionar..."):
            aplicar_favorito(nombre_sel)
            st.rerun()
            
    with f_col3:
        # Botón Eliminar
        if st.button("🗑️ Eliminar", use_container_width=True, help="Eliminar esta configuración", disabled=nombre_sel=="Seleccionar..."):
            if eliminar_favorito(nombre_sel):
                st.toast(f"'{nombre_sel}' eliminado")
                st.rerun()
            
    with f_col4:
        col_txt, col_btn_save = st.columns([1.2, 1])
        with col_txt:
            nuevo_fav = st.text_input("Nombre nuevo", placeholder="Nombre...", label_visibility="collapsed", key="input_nuevo_fav")
        with col_btn_save:
            if st.button("💾 Guardar", use_container_width=True):
                if nuevo_fav:
                    w_glob = {
                        "Ofensivo": st.session_state.get("w_of_val", 0.25), 
                        "Defensivo": st.session_state.get("w_def_val", 0.25), 
                        "Técnico": st.session_state.get("w_tec_val", 0.25), 
                        "Físico": st.session_state.get("w_fis_val", 0.25)
                    }
                    w_mic = {k: v for k, v in st.session_state.items() if k.startswith("slider_")}
                    guardar_favorito(nuevo_fav, w_glob, w_mic)
                    st.toast(f"¡{nuevo_fav} guardado!")
                    st.rerun()
                else:
                    st.warning("Escribe un nombre")
    
    st.divider()

    # --- TUS PREDEFINIDOS ORIGINALES (MANTENIDOS) ---
    st.markdown("<p style='font-weight: bold;'>🎯 Perfiles predefinidos por posición (Excel):</p>", unsafe_allow_html=True)

    posiciones_macro = ["Portero", "Defensa central", "Lateral", "Mediocampista defensivo", 
                        "Mediocampista ofensivo", "Extremo", "Delantero centro"]

    cols_macro = st.columns(4)
    for i, pos_name in enumerate(posiciones_macro):
        with cols_macro[i % 4]:
            if st.button(pos_name, key=f"macro_btn_{pos_name}", use_container_width=True):
                m_vals = cargar_pesos_macro_tabla(str(ruta_pesos_macro), pos_name)
                if m_vals:
                    st.session_state["w_of_val"] = m_vals["Ofensivo"]
                    st.session_state["w_def_val"] = m_vals["Defensivo"]
                    st.session_state["w_tec_val"] = m_vals["Técnico"]
                    st.session_state["w_fis_val"] = m_vals["Físico"]
                    st.rerun()

    st.divider()
    
    # --- SLIDERS DE CONTROL ---
    c_g1, c_g2, c_g3, c_g4 = st.columns(4)
    with c_g1: 
        w_of = st.slider("⚽ Ofensivo", 0.0, 1.0, key="w_of_val", step=0.05)
    with c_g2: 
        w_def = st.slider("🛡️ Defensivo", 0.0, 1.0, key="w_def_val", step=0.05)
    with c_g3: 
        w_tec = st.slider("🪄 Técnico", 0.0, 1.0, key="w_tec_val", step=0.05)
    with c_g4: 
        w_fis = st.slider("🏃 Físico", 0.0, 1.0, key="w_fis_val", step=0.05)

    sum_w = (w_of + w_def + w_tec + w_fis) or 1
    pesos_globales = {
        "Ofensivo": w_of/sum_w, 
        "Defensivo": w_def/sum_w, 
        "Técnico": w_tec/sum_w, 
        "Físico": w_fis/sum_w
    }

# ==================== NIVEL 2: PESOS POR ESTADÍSTICA (MICRO) ====================
with st.expander("🛠️ NIVEL 2: Ajuste del peso individual de cada estadística"):
    st.markdown("<p style='font-weight: bold;'>📊 Cargar pesos por posición (Pestañas Excel):</p>", unsafe_allow_html=True)
    
    hojas_excel = ["Defensa central", "Lateral", "Mediocampista defensivo", "Mediocampista ofensivo", "Extremo", "Delantero"]
    cols_btn = st.columns(3) 

    for i, nombre_hoja in enumerate(hojas_excel):
        with cols_btn[i % 3]:
            if st.button(nombre_hoja, key=f"micro_btn_{nombre_hoja}", use_container_width=True):
                p_vals = cargar_pesos_por_posicion(str(ruta_pesos), nombre_hoja)
                if p_vals:
                    for cat, metricas in p_vals.items():
                        for m, v in metricas.items():
                            st.session_state[f"slider_{cat}_{m}"] = float(v)
                    st.rerun()

    st.divider()
    # Cargamos la estructura inicial para los tabs (usamos Delantero como base estructural)
    PESOS_BASE = cargar_pesos_por_posicion(str(ruta_pesos), "Delantero")
    tabs = st.tabs(["⚽ Ataque", "🛡️ Defensa", "🪄 Técnica", "🏃 Físico"])
    pesos_individuales = {cat: {} for cat in COL_MAP.values()}

    for i, (categoria, metricas) in enumerate(PESOS_BASE.items()):
        with tabs[i]:
            c1, c2 = st.columns(2)
            items = list(metricas.items())
            mid = (len(items) + 1) // 2
            for j, (metrica, valor_def) in enumerate(items):
                key_s = f"slider_{categoria}_{metrica}"
                val_s = st.session_state.get(key_s, float(valor_def))
                with (c1 if j < mid else c2):
                    pesos_individuales[categoria][metrica] = st.slider(metrica, -100.0, 100.0, val_s, 5.0, key=key_s)




def calcular_ranking_total(df_input, pesos_micro, pesos_macro):
    df = df_input.copy()
    
    # 1. Normalización min-max sobre los valores reales (sin tocar los NaN)
    for cat in pesos_micro:
        for m in pesos_micro[cat]:
            if m in df.columns:
                min_v, max_v = df[m].min(), df[m].max()
                if max_v != min_v:
                    df[f"n_{m}"] = (df[m] - min_v) / (max_v - min_v) * 100
                else:
                    df[f"n_{m}"] = 0.0  # todas iguales → 0 diferencia

    # 2. Cálculo del Score del Bloque
    # Para cada jugadora, dividimos solo entre la suma de pesos de métricas con dato.
    # Las que tienen NaN no penalizan ni inflan el score.
    for cat, metricas_pesos in pesos_micro.items():
        activas = {m: w for m, w in metricas_pesos.items() if w != 0 and f"n_{m}" in df.columns}

        if not activas:
            df[f"Score {cat}"] = np.nan
            continue

        cols_norm  = [f"n_{m}" for m in activas]
        pesos_arr  = np.array(list(activas.values()), dtype=float)
        vals       = df[cols_norm].values  # (n_jugadoras, n_metricas), con NaN

        scores = []
        for row in vals:
            mask = ~np.isnan(row)
            if mask.sum() == 0:
                scores.append(np.nan)
            else:
                # Promedio ponderado solo con las métricas disponibles
                scores.append(float(np.dot(row[mask], pesos_arr[mask]) / pesos_arr[mask].sum()))
        df[f"Score {cat}"] = np.round(scores, 1)

    # 3. OVERALL (Nivel Macro)
    df["Score Global"] = (
        df["Score Ofensivo"] * pesos_macro["Ofensivo"] +
        df["Score Defensivo"] * pesos_macro["Defensivo"] +
        df["Score Técnico"] * pesos_macro["Técnico"] +
        df["Score Físico"] * pesos_macro["Físico"]
    ).round(1)
    
    return df

def letra_score(s):
    if pd.isna(s): return "-"
    if s >= 75: return "A"
    if s >= 50: return "B"
    if s >= 25: return "C"
    return "D"

def color_score(s):
    if pd.isna(s): return "#94a3b8"
    if s >= 75: return "#10b981"
    if s >= 50: return "#3b82f6"
    if s >= 25: return "#f59e0b"
    return "#ef4444"

BG_COLOR = {
    "#10b981": "background-color: rgba(16,185,129,0.15)",
    "#3b82f6": "background-color: rgba(59,130,246,0.12)",
    "#f59e0b": "background-color: rgba(245,158,11,0.12)",
    "#ef4444": "background-color: rgba(239,68,68,0.10)",
    "#94a3b8": "background-color: rgba(148,163,184,0.10)",
}

with st.spinner("Calculando scores de rendimiento..."):
    df_scored = calcular_ranking_total(df_filtered, pesos_individuales, pesos_globales)

    # Calcular un df_scored independiente por cada liga con ajuste automático de minutos
    # Si con el mínimo configurado no hay TOP_N_MINIMO jugadoras, se baja de 50 en 50 hasta lograrlo
    TOP_N_MINIMO = 30
    df_scored_por_liga = {}
    df_minutos_ajustados = {}  # registra ligas donde se bajó el mínimo

    for _nombre_arch, _df_liga_raw in archivos_cargados.items():
        _competicion = _competicion_de_nombre(_nombre_arch)
        _df_liga_base = _df_liga_raw.copy()
        # Conversión numérica
        for _col in _df_liga_base.columns:
            if _df_liga_base[_col].dtype == object:
                _converted = _df_liga_base[_col].astype(str).str.replace(",", ".", regex=False)
                _numeric = pd.to_numeric(_converted, errors="coerce")
                if _numeric.notna().sum() > _df_liga_base[_col].notna().sum() * 0.5:
                    _df_liga_base[_col] = _numeric
        if "Posición específica" in _df_liga_base.columns:
            _df_liga_base["Grupo posición"] = _df_liga_base["Posición específica"].apply(get_grupo_posicion)

        # Filtros fijos (no dependen de los minutos)
        def _aplicar_filtros_fijos(df):
            if equipos_sel and "Equipo durante el período seleccionado" in df.columns:
                df = df[df["Equipo durante el período seleccionado"].isin(equipos_sel)]
            if posiciones_sel and "Posición específica" in df.columns:
                _pcodes = [INV_MAP_LABELS.get(lbl, lbl) for lbl in posiciones_sel]
                df = df[df["Posición específica"].apply(
                    lambda cell: any(p in [x.strip() for x in str(cell).split(",")]
                                     for p in _pcodes) if pd.notna(cell) else False
                )]
            if grupos_sel and "Grupo posición" in df.columns:
                df = df[df["Grupo posición"].isin(grupos_sel)]
            if edad_activo and rango_edad is not None and "Edad" in df.columns:
                df = df[df["Edad"].between(rango_edad[0], rango_edad[1])]
            return df

        _df_liga_fijos = _aplicar_filtros_fijos(_df_liga_base)

        # Ajuste adaptativo del mínimo de minutos
        _min_min_original = rango_minutos[0]
        _max_min = rango_minutos[1]
        _min_min_usado = _min_min_original
        _df_liga = pd.DataFrame()

        if "Minutos jugados" in _df_liga_fijos.columns:
            _umbral = _min_min_original
            while True:
                _df_candidata = _df_liga_fijos[
                    _df_liga_fijos["Minutos jugados"].between(_umbral, _max_min)
                ]
                if len(_df_candidata) >= TOP_N_MINIMO or _umbral == 0:
                    _df_liga = _df_candidata
                    _min_min_usado = _umbral
                    break
                _umbral = max(0, _umbral - 50)
        else:
            _df_liga = _df_liga_fijos

        if _df_liga.empty:
            continue

        # Registrar ajuste aplicado
        if _min_min_usado < _min_min_original:
            df_minutos_ajustados[_competicion] = (_min_min_original, _min_min_usado, len(_df_liga))

        _df_liga = add_per90_cols(_df_liga)
        _df_sc = calcular_ranking_total(_df_liga, pesos_individuales, pesos_globales)
        # Calcular z-score y percentil dentro de esta liga
        for _cat in ["Global", "Ofensivo", "Defensivo", "Técnico", "Físico"]:
            _sc_col = f"Score {_cat}"
            if _sc_col in _df_sc.columns:
                _vals = _df_sc[_sc_col].dropna()
                if not _vals.empty:
                    _df_sc[f"Z-score {_cat}"] = zscore(_df_sc[_sc_col], nan_policy="omit").round(2)
                    _df_sc[f"Percentil {_cat}"] = [
                        round(percentileofscore(_vals, v, kind="rank"), 1) if pd.notna(v) else np.nan
                        for v in _df_sc[_sc_col]
                    ]
                else:
                    _df_sc[f"Z-score {_cat}"] = np.nan
                    _df_sc[f"Percentil {_cat}"] = np.nan
        df_scored_por_liga[_competicion] = _df_sc

    # Ranking global: unión de todas las ligas con percentiles recalculados sobre el total
    _KEY_TODAS = "🌍 Todas las ligas"
    if len(df_scored_por_liga) > 1:
        _dfs_union = []
        for _liga_key, _df_l in df_scored_por_liga.items():
            _tmp = _df_l.copy()
            _tmp["_liga"] = _liga_key
            _dfs_union.append(_tmp)
        _df_todas = pd.concat(_dfs_union, ignore_index=True)
        # Recalcular z-score y percentil sobre el conjunto completo
        for _cat in ["Global", "Ofensivo", "Defensivo", "Técnico", "Físico"]:
            _sc_col = f"Score {_cat}"
            if _sc_col in _df_todas.columns:
                _vals = _df_todas[_sc_col].dropna()
                if not _vals.empty:
                    _df_todas[f"Z-score {_cat}"] = zscore(_df_todas[_sc_col], nan_policy="omit").round(2)
                    _df_todas[f"Percentil {_cat}"] = [
                        round(percentileofscore(_vals, v, kind="rank"), 1) if pd.notna(v) else np.nan
                        for v in _df_todas[_sc_col]
                    ]
                else:
                    _df_todas[f"Z-score {_cat}"] = np.nan
                    _df_todas[f"Percentil {_cat}"] = np.nan
        df_scored_por_liga = {_KEY_TODAS: _df_todas, **df_scored_por_liga}

score_cols_all = ["Score Ofensivo", "Score Defensivo", "Score Técnico", "Score Físico", "Score Global"]

# Columnas de identidad (sobre df_scored global, válido para detectar nombres de columna)
col_nombre  = next((c for c in ["Jugador", "player_name", "Player Name"] if c in df_scored.columns), None)
col_equipo  = next((c for c in ["Equipo durante el período seleccionado"]        if c in df_scored.columns), None)
col_edad  = next((c for c in ["Edad", "Age", "Birthday"]        if c in df_scored.columns), None)
col_posicion= next((c for c in ["Posición específica", "position"]    if c in df_scored.columns), None)
col_minutos = "Minutos jugados" if "Minutos jugados" in df_scored.columns else None
col_partidos = "Partidos jugados" if "Partidos jugados" in df_scored.columns else None

tab1, tab2, tab3 = st.tabs(["📊 Rankings de Rendimiento", "🔍 Visualizador por Métrica", "📋 Tabla Multi-Métrica"])

with tab1:
    # ==================== TABLA GENERAL ====================
    st.markdown("<h2 class='section-header'>📊 Rankings de Rendimiento</h2>", unsafe_allow_html=True)

    st.markdown("""
        <div class='info-box' style='background: linear-gradient(135deg, rgba(59,130,246,0.05) 0%, rgba(118,75,162,0.05) 100%); border-left: 4px solid #3b82f6;'>
            <p style='margin: 0; font-size: 0.85rem; line-height: 1.6; color: #475569;'>
                Selecciona una categoría para ver el desglose detallado. Cada tabla incluye <b>Puntuación</b>, <b>z-score</b> y <b>Percentil</b> calculado sobre el grupo filtrado.
            </p>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- AVISO DE AJUSTE AUTOMÁTICO DE MINUTOS ---
    if df_minutos_ajustados:
        lineas = []
        for _liga, (_orig, _usado, _n) in df_minutos_ajustados.items():
            lineas.append(f"<b>{_liga}</b>: mínimo bajado de {_orig} a {_usado} min → {_n} jugadoras")
        st.warning(
            "⚠️ **Ajuste automático de minutos** — Las siguientes ligas no alcanzaban 30 jugadoras "
            "con el filtro configurado y se redujo el mínimo automáticamente:\n\n" +
            "\n\n".join(f"• {l}" for l in lineas),
            icon=None
        )

    # --- CONTROLES DE LA TABLA ---
    max_jugadoras = int(df_scored["Score Global"].notna().sum()) if "Score Global" in df_scored.columns else 100
    
    ctrl_a, _, ctrl_b = st.columns([1, 0.5, 1])

    with ctrl_a:
        top_n_tabla = st.slider("Top N jugadoras", 1, max_jugadoras, min(30, max_jugadoras), 1, key="top_n_rank")

    with ctrl_b:
        st.markdown("Filtrar por Nota")
        n1, n2, n3, n4 = st.columns(4)
        with n1: chk_a = st.checkbox("A", value=True, key="rank_a")
        with n2: chk_b = st.checkbox("B", value=True, key="rank_b")
        with n3: chk_c = st.checkbox("C", value=True, key="rank_c")
        with n4: chk_d = st.checkbox("D", value=True, key="rank_d")
        notas_sel = [n for n, chk in [("A", chk_a), ("B", chk_b), ("C", chk_c), ("D", chk_d)] if chk]

    st.markdown("<br>", unsafe_allow_html=True)

    # --- INTERFAZ DE NAVEGACIÓN POR PESTAÑAS DE CATEGORÍA ---
    tab_overall, tab_ofensivo, tab_defensivo, tab_tecnico, tab_fisico = st.tabs([
        "🏆 GLOBAL", "⚽ OFENSIVO", "🛡️ DEFENSIVO", "🪄 TÉCNICO", "🏃 FÍSICO"
    ])

    def renderizar_ranking_liga(df_sc_liga, categoria_id, liga_key):
        """Renderiza la tabla de ranking para una liga y categoría concretas."""
        sc = f"Score {categoria_id}"
        zs = f"Z-score {categoria_id}"
        pr = f"Percentil {categoria_id}"
        nt = f"Nota {categoria_id}"

        base_cols = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, col_minutos] if c]
        available = [c for c in base_cols + [sc, zs, pr] if c in df_sc_liga.columns]
        df_cat = df_sc_liga[available].copy().dropna(subset=[sc] if sc in available else [])

        if pr in df_cat.columns:
            df_cat = df_cat[df_cat[pr].apply(letra_score).isin(notas_sel)]

        df_cat = df_cat.sort_values(sc, ascending=False)

        if nt not in df_cat.columns and pr in df_cat.columns:
            df_cat.insert(df_cat.columns.get_loc(pr) + 1, nt, df_cat[pr].apply(letra_score))

        rename_map = {}
        if col_nombre:   rename_map[col_nombre]   = "Jugadora"
        if col_equipo:   rename_map[col_equipo]   = "Equipo"
        if col_edad:     rename_map[col_edad]     = "Edad"
        if col_posicion: rename_map[col_posicion] = "Posición"
        if col_minutos:  rename_map[col_minutos]  = "Min"
        df_cat = df_cat.rename(columns=rename_map).head(top_n_tabla)

        busqueda = st.text_input(
            f"🔍 Buscar jugadora",
            placeholder="Escribe un nombre...",
            key=f"buscar_{liga_key}_{categoria_id}"
        )
        df_cat = df_cat.reset_index(drop=True)
        if busqueda and "Jugadora" in df_cat.columns:
            mask = df_cat["Jugadora"].str.contains(busqueda, case=False, na=False)
            df_cat = pd.concat([df_cat[mask], df_cat[~mask]]).reset_index(drop=True)

        sel_key = f"sel_nombres_{liga_key}_{categoria_id}"
        nombres_sel_tabla = st.session_state.get(sel_key, set())
        df_cat.insert(0, "✓", df_cat["Jugadora"].isin(nombres_sel_tabla) if "Jugadora" in df_cat.columns else False)

        fmt = {}
        if sc in df_cat.columns: fmt[sc] = "{:.1f}"
        if pr in df_cat.columns: fmt[pr] = "{:.1f}"
        if zs in df_cat.columns: fmt[zs] = "{:.2f}"
        if "Edad" in df_cat.columns: fmt["Edad"] = "{:.0f}"
        if "Min" in df_cat.columns: fmt["Min"] = "{:.0f}"

        gradient_subset = [sc] if sc in df_cat.columns else []
        edited = st.data_editor(
            df_cat.style.format(fmt, na_rep="-").background_gradient(subset=gradient_subset, cmap="RdYlGn"),
            use_container_width=True,
            height=450,
            key=f"grid_{liga_key}_{categoria_id}",
            column_config={
                "✓": st.column_config.CheckboxColumn("✓", help="Marcar para descargar", width="small"),
                **({sc: st.column_config.NumberColumn(sc, format="%.1f")} if sc in df_cat.columns else {}),
                **({pr: st.column_config.NumberColumn(pr, format="%.1f")} if pr in df_cat.columns else {}),
                **({zs: st.column_config.NumberColumn(zs, format="%.2f")} if zs in df_cat.columns else {}),
            },
            disabled=[c for c in df_cat.columns if c != "✓"],
        )

        nuevos_sel = set(edited.loc[edited["✓"] == True, "Jugadora"].tolist()) if "Jugadora" in edited.columns else set()
        st.session_state[sel_key] = nuevos_sel

    @st.fragment
    def renderizar_categoria(categoria_id):
        """Renderiza todas las ligas para una categoría dada."""
        _KEY_TODAS = "🌍 Todas las ligas"
        ligas_disponibles = list(df_scored_por_liga.keys())

        if len(ligas_disponibles) == 1:
            # Solo una liga: comportamiento original
            liga = ligas_disponibles[0]
            renderizar_ranking_liga(df_scored_por_liga[liga], categoria_id, liga_key=liga)
        else:
            # Sub-tab global primero, luego una por liga
            tab_labels = [liga if liga == _KEY_TODAS else f"🏟️ {liga}" for liga in ligas_disponibles]
            sub_tabs = st.tabs(tab_labels)
            for sub_tab, liga in zip(sub_tabs, ligas_disponibles):
                with sub_tab:
                    renderizar_ranking_liga(df_scored_por_liga[liga], categoria_id, liga_key=liga)

    # Renderizado de fragmentos dentro de sus respectivos tabs
    with tab_overall:
        renderizar_categoria("Global")
    with tab_ofensivo:
        renderizar_categoria("Ofensivo")
    with tab_defensivo:
        renderizar_categoria("Defensivo")
    with tab_tecnico:
        renderizar_categoria("Técnico")
    with tab_fisico:
        renderizar_categoria("Físico")


    # ── Botón de descarga: ZIP con un Excel por liga, 5 hojas cada uno ────────
    st.info("💡 Selecciona filas en las tablas superiores para incluir solo esas jugadoras en la descarga.")

    def _preparar_hoja(df_scored_liga, categoria_id: str, liga_key: str = "") -> pd.DataFrame:
        sc = f"Score {categoria_id}"
        zs = f"Z-score {categoria_id}"
        pr = f"Percentil {categoria_id}"
        nt = f"Nota {categoria_id}"
        base_cols = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, col_minutos] if c]
        available = [c for c in base_cols + [sc, zs, pr] if c in df_scored_liga.columns]
        df_cat = df_scored_liga[available].copy().dropna(subset=[sc] if sc in available else [])
        if pr in df_cat.columns:
            df_cat = df_cat[df_cat[pr].apply(letra_score).isin(notas_sel)]
        if sc in df_cat.columns:
            df_cat = df_cat.sort_values(sc, ascending=False)
        if nt not in df_cat.columns and pr in df_cat.columns:
            df_cat.insert(df_cat.columns.get_loc(pr) + 1, nt, df_cat[pr].apply(letra_score))
        rename_map = {}
        if col_nombre:   rename_map[col_nombre]   = "Jugadora"
        if col_equipo:   rename_map[col_equipo]   = "Equipo"
        if col_edad:     rename_map[col_edad]     = "Edad"
        if col_posicion: rename_map[col_posicion] = "Posición"
        if col_minutos:  rename_map[col_minutos]  = "Min"
        df_cat = df_cat.rename(columns=rename_map).head(top_n_tabla).reset_index(drop=True)
        # Usar clave de selección por liga si está disponible, si no caer al formato antiguo
        sel_key = f"sel_nombres_{liga_key}_{categoria_id}" if liga_key else f"sel_nombres_{categoria_id}"
        nombres_sel_cat = st.session_state.get(sel_key, set())
        if nombres_sel_cat and "Jugadora" in df_cat.columns:
            df_cat = df_cat[df_cat["Jugadora"].isin(nombres_sel_cat)].reset_index(drop=True)
        return df_cat

    def _escribir_hoja_excel(wb, df, sheet_title):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
        ws = wb.create_sheet(title=sheet_title[:31])
        header_fill  = PatternFill("solid", fgColor="1E3A8A")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        row_fill_alt = PatternFill("solid", fgColor="F1F5F9")
        row_fill_wht = PatternFill("solid", fgColor="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left",   vertical="center")
        thin_side    = Side(style="thin", color="CBD5E1")
        thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        text_cols    = {"Jugadora", "Equipo", "Posición"}
        score_kw     = ("Score", "Percentil", "Z-score")
        headers = ["#"] + list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center_align; cell.border = thin_border
        ws.row_dimensions[1].height = 22
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            fill = row_fill_alt if row_idx % 2 == 0 else row_fill_wht
            idx_cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            idx_cell.fill = fill; idx_cell.font = Font(size=10, color="64748B")
            idx_cell.alignment = center_align; idx_cell.border = thin_border
            for col_idx, (col_name, value) in enumerate(row.items(), start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill; cell.border = thin_border
                cell.alignment = left_align if (col_name in text_cols or isinstance(value, str)) else center_align
                cell.font = Font(size=10)
                if isinstance(value, float): cell.number_format = "0.00"
                elif isinstance(value, int): cell.number_format = "0"
        last_row = ws.max_row
        for col_idx, col_name in enumerate(headers, start=1):
            if any(kw in str(col_name) for kw in score_kw) and last_row > 2:
                col_letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{last_row}",
                    ColorScaleRule(start_type="min", start_color="EF4444",
                                   mid_type="percentile", mid_value=50, mid_color="F59E0B",
                                   end_type="max", end_color="10B981")
                )
        for col_idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)),
                          *[len(str(ws.cell(row=r, column=col_idx).value or ""))
                            for r in range(2, last_row + 1)])
            ws.column_dimensions[col_letter].width = min(max_len + 3, 35)
        ws.freeze_panes = "A2"

    def _generar_zip_ligas() -> bytes:
        import zipfile
        from openpyxl import Workbook

        categorias_hojas = [
            ("Global",    "🏆 Global"),
            ("Ofensivo",  "⚽ Ofensivo"),
            ("Defensivo", "🛡️ Defensivo"),
            ("Técnico",   "🪄 Técnico"),
            ("Físico",    "🏃 Físico"),
        ]
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for competicion, df_scored_liga in df_scored_por_liga.items():
                # Crear Excel con 5 hojas usando los datos ya calculados
                wb = Workbook()
                wb.remove(wb.active)
                for cat_id, sheet_title in categorias_hojas:
                    df_hoja = _preparar_hoja(df_scored_liga, cat_id, liga_key=competicion)
                    _escribir_hoja_excel(wb, df_hoja, sheet_title)
                excel_buf = io.BytesIO()
                wb.save(excel_buf)
                _KEY_TODAS = "🌍 Todas las ligas"
                nombre_archivo_zip = (
                    f"ranking_todas_las_ligas_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
                    if competicion == _KEY_TODAS
                    else _nombre_descarga(competicion=competicion)
                )
                zf.writestr(nombre_archivo_zip, excel_buf.getvalue())
        return zip_buf.getvalue()

    # Nombre del ZIP y label del botón
    pos_part = (
        "-".join(INV_MAP_LABELS.get(lbl, lbl) for lbl in posiciones_sel).replace(" ", "")
        if posiciones_sel else "todas"
    )
    n_ligas = len(df_scored_por_liga)
    nombre_zip = (
        f"rankings_{pos_part}.zip" if n_ligas > 1
        else _nombre_descarga(competicion=list(df_scored_por_liga.keys())[0]).replace(".xlsx", ".zip")
    )
    total_sel = sum(
        len(st.session_state.get(f"sel_nombres_{liga}_{cat}", set()))
        for liga in df_scored_por_liga
        for cat in ["Global", "Ofensivo", "Defensivo", "Técnico", "Físico"]
    )
    if n_ligas > 1:
        label_descarga = (
            f"📥 Descargar rankings — {n_ligas} Excel independientes (1 por liga · 5 hojas cada uno)"
            if not total_sel else
            f"📥 Descargar rankings — {n_ligas} ligas · {total_sel} jugadoras seleccionadas"
        )
    else:
        label_descarga = (
            f"📥 Descargar rankings · 5 hojas"
            if not total_sel else
            f"📥 Descargar rankings ({total_sel} jugadoras seleccionadas)"
        )
    st.download_button(
        label=label_descarga,
        data=_generar_zip_ligas(),
        file_name=nombre_zip,
        mime="application/zip",
        key="dl_ranking_zip",
        type="primary",
        use_container_width=True,
    )
    st.divider()

with tab2:
    @st.fragment
    def render_tab2():
        # ==================== ANÁLISIS POR MÉTRICA INDIVIDUAL ====================
        st.markdown("<h2 class='section-header'>🔍 Visualizador por Métrica Individual</h2>", unsafe_allow_html=True)

        fil1, fil2, fil3, fil4 = st.columns([2, 2, 1.5, 1.5])

        with fil1:
            categoria_metrica = st.selectbox(
                "Categoría de métricas",
                ["Ofensivo", "Defensivo", "Técnico", "Físico"],
                key="categoria_metrica"
            )
        with fil2:
            tipo_viz = st.radio(
                "Tipo de visualización",
                ["Tabla detallada", "Dispersión (scatter)"],
                key="tipo_viz"
            )
        with fil3:
            top_n = st.slider("Top N", 5, max_jugadoras, max_jugadoras, 5, key="top_n_ind")
        with fil4:
            orden = st.radio("Orden", ["Mayor a menor ↓", "Menor a mayor ↑"], key="orden_ranking")

        st.markdown("<br>", unsafe_allow_html=True)

        METRICAS_CATEGORIAS = {
            cat: {var: var for var in vars_dict}
            for cat, vars_dict in pesos_individuales.items()
        }
        metricas_ya_en_cats = {m for cat in METRICAS_CATEGORIAS.values() for m in cat}
        metricas_extra = {
            nombre: col
            for nombre, col in RENAME_MAP_METRICAS.items()
            if nombre not in metricas_ya_en_cats and col in df_scored.columns
        }
        for cat in METRICAS_CATEGORIAS:
            METRICAS_CATEGORIAS[cat].update(metricas_extra)
        METRICAS_CATEGORIAS["Global"] = {
            "Score Global":   "Score Global",
            "Score Ofensivo": "Score Ofensivo",
            "Score Defensivo":"Score Defensivo",
            "Score Técnico":  "Score Técnico",
            "Score Físico":   "Score Físico",
        }

        metricas_categoria = METRICAS_CATEGORIAS.get(categoria_metrica, {})
        metricas_disponibles = {
            nombre: col
            for nombre, col in metricas_categoria.items()
            if col in df_scored.columns
        }

        if not metricas_disponibles:
            st.warning(f"⚠️ No se encontraron métricas para '{categoria_metrica}'.")
            return

        sel1, sel2 = st.columns([3, 1])
        with sel1:
            metrica_sel_nombre = st.selectbox(
                f"Métrica ({categoria_metrica})", list(metricas_disponibles.keys()), key="metrica_sel"
            )
        with sel2:
            tipo_ranking = st.radio("Ranking por", ["Valor", "Z-score", "Percentil"], key="tipo_ranking")

        metrica_col = metricas_disponibles[metrica_sel_nombre]
        ascending = orden == "Menor a mayor ↑"

        ranking_cols = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, col_partidos, col_minutos, metrica_col] if c]
        df_ranking = df_scored[ranking_cols].copy().dropna(subset=[metrica_col])

        vals = df_ranking[metrica_col].values
        df_ranking["Z-score"]   = zscore(vals, ddof=1).round(2)
        df_ranking["Percentil"] = [round(percentileofscore(vals, v, kind="rank"), 1) for v in vals]

        col_sort = {"Valor": metrica_col, "Z-score": "Z-score", "Percentil": "Percentil"}[tipo_ranking]
        df_ranking = df_ranking.sort_values(col_sort, ascending=ascending)

        if len(notas_sel) < 4 and "Percentil Global" in df_scored.columns:
            df_ranking_tmp = df_scored[ranking_cols].copy().dropna(subset=[metrica_col])
            df_ranking_tmp["_nota_global"] = df_scored.loc[df_ranking_tmp.index, "Percentil Global"].apply(letra_score)
            df_ranking_tmp = df_ranking_tmp[df_ranking_tmp["_nota_global"].isin(notas_sel)]
            df_ranking = df_ranking_tmp.drop(columns=["_nota_global"])
            df_ranking["Z-score"] = zscore(df_ranking[metrica_col].values, ddof=1).round(2)
            df_ranking["Percentil"] = [round(percentileofscore(df_ranking[metrica_col].values, v, kind="rank"), 1) for v in df_ranking[metrica_col].values]
            df_ranking = df_ranking.sort_values(col_sort, ascending=ascending)

        df_ranking_top = df_ranking.head(top_n).reset_index(drop=True)
        df_ranking_top.index += 1

        rename_map = {}
        if col_nombre:   rename_map[col_nombre]  = "Jugadora"
        if col_equipo:   rename_map[col_equipo]  = "Equipo"
        if col_edad:     rename_map[col_edad]    = "Edad"
        if col_posicion: rename_map[col_posicion]= "Posición"
        if col_minutos:  rename_map[col_minutos] = "Minutos"
        if col_partidos: rename_map[col_partidos]= "Partidos Jugados"
        rename_map[metrica_col] = metrica_sel_nombre

        df_display = df_ranking_top.rename(columns=rename_map)

        info1, info2, info3, info4 = st.columns(4)
        with info1: st.metric("Jugadoras analizadas", len(df_ranking))
        with info2: st.metric(f"Máx ({tipo_ranking})", f"{df_ranking[col_sort].max():.2f}")
        with info3: st.metric(f"Media ({tipo_ranking})", f"{df_ranking[col_sort].mean():.2f}")
        with info4: st.metric(f"Mín ({tipo_ranking})", f"{df_ranking[col_sort].min():.2f}")

        st.markdown("<br>", unsafe_allow_html=True)

        # Buscador dentro del tab
        busqueda_ind = st.text_input(
            "🔍 Buscar jugadora",
            placeholder="Escribe un nombre...",
            key="buscar_ind"
        )

        # ─── TABLA DETALLADA ───
        if tipo_viz == "Tabla detallada":
            tabla_cols = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, col_partidos, col_minutos, metrica_col, "Z-score", "Percentil"] if c]
            df_tabla = df_ranking_top[tabla_cols].copy().rename(columns={**rename_map, "Z-score": "Z-score", "Percentil": "Percentil"})

            fmt = {metrica_sel_nombre: "{:.2f}", "Z-score": "{:.2f}", "Percentil": "{:.1f}"}
            if "Minutos" in df_tabla.columns: fmt["Minutos"] = "{:.0f}"
            if "Edad" in df_tabla.columns: fmt["Edad"] = "{:.0f}"

            df_tabla = df_tabla.reset_index(drop=True)
            if busqueda_ind and "Jugadora" in df_tabla.columns:
                mask = df_tabla["Jugadora"].str.contains(busqueda_ind, case=False, na=False)
                df_tabla = pd.concat([df_tabla[mask], df_tabla[~mask]]).reset_index(drop=True)

            nombres_sel_ind = st.session_state.get("sel_nombres_ind", set())
            df_tabla.insert(0, "✓", df_tabla["Jugadora"].isin(nombres_sel_ind))

            edited_ind = st.data_editor(
                df_tabla.style.format(fmt, na_rep="-"),
                use_container_width=True,
                height=600,
                key="grid_individual",
                column_config={
                    "✓": st.column_config.CheckboxColumn("✓", help="Marcar para descargar", width="small"),
                },
                disabled=[c for c in df_tabla.columns if c != "✓"],
            )

            nuevos_sel_ind = set(edited_ind.loc[edited_ind["✓"] == True, "Jugadora"].tolist())
            st.session_state["sel_nombres_ind"] = nuevos_sel_ind

            df_dl_ind = df_tabla[df_tabla["Jugadora"].isin(nuevos_sel_ind)].drop(columns=["✓"]) if nuevos_sel_ind else df_tabla.drop(columns=["✓"])
            txt_dl = f"📊 Descargar {len(nuevos_sel_ind)} jugadoras seleccionadas" if nuevos_sel_ind else "📊 Descargar Top completo"

            excel_bytes_ind = exportar_excel_bonito(df_dl_ind.reset_index(drop=True), sheet_name="Detalle")
            st.download_button(
                label=txt_dl,
                data=excel_bytes_ind,
                file_name=_nombre_descarga(metrica_sel_nombre.replace(" ", "_")),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_ind_btn"
            )

        # ─── SCATTER ───
        elif tipo_viz == "Dispersión (scatter)":
            metricas_x = {k: v for k, v in metricas_disponibles.items() if k != metrica_sel_nombre}
            if not metricas_x:
                st.info("Selecciona una categoría con al menos 2 métricas para usar scatter.")
            else:
                metrica_x_nombre = st.selectbox("Métrica eje X", list(metricas_x.keys()), key="metrica_x_scatter")
                metrica_x_col = metricas_x[metrica_x_nombre]

                scatter_cols = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, metrica_col, metrica_x_col] if c]
                df_scatter = df_scored[scatter_cols].dropna().copy()
                vals_sc = df_scatter[metrica_col].values
                df_scatter["Z-score"]   = zscore(vals_sc, ddof=1).round(2)
                df_scatter["Percentil"] = [round(percentileofscore(vals_sc, v, kind="rank"), 1) for v in vals_sc]

                color_col = col_posicion if (col_posicion and not posiciones_sel) else col_equipo
                fig = px.scatter(
                    df_scatter, x=metrica_x_col, y=metrica_col, color=color_col,
                    hover_name=col_nombre if col_nombre else None,
                    hover_data={metrica_col: ":.2f", metrica_x_col: ":.2f", "Z-score": ":.2f", "Percentil": ":.1f"},
                    labels={metrica_col: metrica_sel_nombre, metrica_x_col: metrica_x_nombre},
                    title=f"{metrica_sel_nombre} vs {metrica_x_nombre}",
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                avg_y, avg_x = df_scatter[metrica_col].mean(), df_scatter[metrica_x_col].mean()
                fig.add_hline(y=avg_y, line_dash="dash", line_color="#94a3b8",
                              annotation_text=f"Promedio: {avg_y:.2f}", annotation_position="top right")
                fig.add_vline(x=avg_x, line_dash="dash", line_color="#94a3b8",
                              annotation_text=f"Promedio: {avg_x:.2f}", annotation_position="top right")
                fig.update_layout(height=550, paper_bgcolor='white', plot_bgcolor='#f8fafc',
                                  font=dict(color="#475569"), title=dict(font=dict(size=18, color="#1e3a8a"), x=0.5))
                fig.update_traces(marker=dict(size=9, opacity=0.8))
                st.plotly_chart(fig, use_container_width=True)

        st.divider()

    render_tab2()

with tab3:
    @st.fragment
    def render_tab3():
        # ==================== TABLA GLOBAL MULTI-MÉTRICA ====================
        st.markdown("<h2 class='section-header'>📋 Tabla Global Multi-Métrica</h2>", unsafe_allow_html=True)

        todas_metricas = {
            nombre: col
            for nombre, col in RENAME_MAP_METRICAS.items()
            if col in df_scored.columns
        }
        for sc_col in ["Score Global", "Score Ofensivo", "Score Defensivo", "Score Técnico", "Score Físico"]:
            if sc_col in df_scored.columns:
                todas_metricas[sc_col] = sc_col

        col_met_sel, = st.columns([1])
        with col_met_sel:
            metricas_tabla_sel = st.multiselect(
                "Selecciona métricas para la tabla global",
                list(todas_metricas.keys()),
                default=list(todas_metricas.keys())[:6] if todas_metricas else [],
                key="metricas_tabla_global"
            )

        if metricas_tabla_sel:
            primera_col = todas_metricas[metricas_tabla_sel[0]]
            cols_tabla = [c for c in [col_nombre, col_equipo, col_edad, col_posicion, col_partidos, col_minutos] if c] + \
                         [todas_metricas[m] for m in metricas_tabla_sel]
            df_tabla_global = df_scored[cols_tabla].copy().dropna(subset=[primera_col])

            if len(notas_sel) < 4 and "Percentil Global" in df_scored.columns:
                nota_global_multi = df_scored.loc[df_tabla_global.index, "Percentil Global"].apply(letra_score)
                df_tabla_global = df_tabla_global[nota_global_multi.isin(notas_sel).values]

            df_tabla_global = df_tabla_global.sort_values(primera_col, ascending=False)
            df_tabla_global = df_tabla_global.head(top_n_tabla)
            df_tabla_global = df_tabla_global.reset_index(drop=True)
            df_tabla_global.index += 1

            rg = {}
            if col_nombre:   rg[col_nombre]  = "Jugadora"
            if col_equipo:   rg[col_equipo]  = "Equipo"
            if col_edad:     rg[col_edad]    = "Edad"
            if col_posicion: rg[col_posicion]= "Posición"
            if col_minutos:  rg[col_minutos] = "Minutos"
            if col_partidos: rg[col_partidos]= "Partidos Jugados"
            for m in metricas_tabla_sel:
                rg[todas_metricas[m]] = m
            df_tabla_global = df_tabla_global.rename(columns=rg)

            numeric_cols = [m for m in metricas_tabla_sel if m in df_tabla_global.columns]
            fmt_g = {c: "{:.2f}" for c in numeric_cols}
            if "Edad" in df_tabla_global.columns: fmt_g["Edad"] = "{:.0f}"

            busqueda_multi = st.text_input(
                "🔍 Buscar jugadora",
                placeholder="Escribe un nombre...",
                key="buscar_multi"
            )

            df_tabla_global = df_tabla_global.reset_index(drop=True)
            if busqueda_multi and "Jugadora" in df_tabla_global.columns:
                mask = df_tabla_global["Jugadora"].str.contains(busqueda_multi, case=False, na=False)
                df_tabla_global = pd.concat([df_tabla_global[mask], df_tabla_global[~mask]]).reset_index(drop=True)

            nombres_sel_multi = st.session_state.get("sel_nombres_multi", set())
            df_tabla_global.insert(0, "✓", df_tabla_global["Jugadora"].isin(nombres_sel_multi))

            edited_multi = st.data_editor(
                df_tabla_global.style.format(fmt_g, na_rep="-").background_gradient(
                    subset=[numeric_cols[0]] if numeric_cols else [], cmap="RdYlGn"
                ),
                use_container_width=True,
                height=500,
                key="grid_multi",
                column_config={
                    "✓": st.column_config.CheckboxColumn("✓", help="Marcar para descargar", width="small"),
                },
                disabled=[c for c in df_tabla_global.columns if c != "✓"],
            )

            nuevos_sel_multi = set(edited_multi.loc[edited_multi["✓"] == True, "Jugadora"].tolist())
            st.session_state["sel_nombres_multi"] = nuevos_sel_multi

            df_dl_m = df_tabla_global[df_tabla_global["Jugadora"].isin(nuevos_sel_multi)].drop(columns=["✓"]) if nuevos_sel_multi else df_tabla_global.drop(columns=["✓"])
            txt_dl = f"📥 Descargar {len(nuevos_sel_multi)} seleccionadas" if nuevos_sel_multi else "📥 Descargar tabla completa"

            excel_bytes_m = exportar_excel_bonito(df_dl_m.reset_index(drop=True), sheet_name="Métricas_Comparadas")
            st.download_button(
                txt_dl, excel_bytes_m, _nombre_descarga("comparativa"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_multi_btn"
            )
        else:
            st.info("Selecciona al menos una métrica para generar la tabla global.")

        st.divider()

    render_tab3()

# ==================== FOOTER ====================
st.markdown(f"""
    <div style='text-align: center; padding: 2rem 0; color: #94a3b8; font-size: 0.85rem;'>
        <p style='margin: 0;'>Liga F Analytics | Temporada {season}</p>
        <p style='margin: 0.5rem 0 0 0;'>Última actualización: {datetime.now().strftime('%d/%m/%Y')}</p>
    </div>
""", unsafe_allow_html=True)
